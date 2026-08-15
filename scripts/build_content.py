#!/usr/bin/env python3
"""
build_content.py  —  Northia website copy compiler.

Edit the wording in  content.xlsx  (any of the sheets), then run this script
to regenerate  content.json , which the website reads at load time.

    python scripts/build_content.py

Then commit the updated content.json (and content.xlsx) to GitHub. That's it —
no HTML/CSS/JS editing required to change wording.

Requires: openpyxl   ->   pip install openpyxl
"""
import json, sys, os
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
XLSX = os.path.join(ROOT, "content.xlsx")
OUT  = os.path.join(ROOT, "content.json")

UI_HEADERS      = ['Key','Where it appears','Variant','EN','繁體中文','简体中文','Notes']
SVC_HEADERS     = ['Ref','EN Name','繁體 Name','简体 Name','EN Description','繁體 Description','简体 Description']
LOC_HEADERS     = ['id','x','y','label_side','flag','category',
                   'EN Name','繁體 Name','简体 Name','EN Role','繁體 Role','简体 Role',
                   'EN Description','繁體 Description','简体 Description',
                   'Article EN title','Article 繁體 title','Article 简体 title','Article slug',
                   'Article body EN','Article body 繁體','Article body 简体']
PARTNER_HEADERS = ['Partner name']

def s(v):
    return "" if v is None else str(v)

def rows(ws):
    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    for r in it:
        if r is None:
            continue
        if all(c is None or str(c).strip()=="" for c in r):
            continue
        yield r

def build_i18n(wb):
    ws = wb["UI Copy"]
    i18n = {}
    for r in rows(ws):
        key = s(r[0]).strip()
        if not key:
            continue
        variant = s(r[2]).strip()
        cell = {"en": s(r[3]), "tc": s(r[4]), "sc": s(r[5])}
        if variant in ("a2w", "w2a"):
            entry = i18n.setdefault(key, {})
            entry[variant] = cell
        else:
            i18n[key] = cell
    return i18n

def build_services(wb):
    ws = wb["Services"]
    out = []
    for r in rows(ws):
        out.append({
            "n": s(r[0]).strip(),
            "name": {"en": s(r[1]), "tc": s(r[2]), "sc": s(r[3])},
            "d":    {"en": s(r[4]), "tc": s(r[5]), "sc": s(r[6])},
        })
    return out

def num(v, default=0):
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return float(v)
        except (TypeError, ValueError):
            return default

def build_locations(wb):
    ws = wb["Locations"]
    out = []
    for r in rows(ws):
        loc = {
            "id": s(r[0]).strip(),
            "x": num(r[1]), "y": num(r[2]), "ls": num(r[3]),
            "flag": s(r[4]),
            "cat": s(r[5]).strip(),
            "name": {"en": s(r[6]), "tc": s(r[7]), "sc": s(r[8])},
            "role": {"en": s(r[9]), "tc": s(r[10]), "sc": s(r[11])},
            "desc": {"en": s(r[12]), "tc": s(r[13]), "sc": s(r[14])},
            "arts": [],
        }
        # One featured article with full body (title + slug + body)
        if s(r[15]).strip() or s(r[18]).strip():
            loc["arts"].append({
                "t": {"en": s(r[15]), "tc": s(r[16]), "sc": s(r[17])},
                "s": s(r[18]).strip(),
                "body": {"en": s(r[19]), "tc": s(r[20]), "sc": s(r[21])},
            })
        out.append(loc)
    return out

def build_brands(wb):
    ws = wb["Partners"]
    return [s(r[0]).strip() for r in rows(ws) if s(r[0]).strip()]

def main():
    if not os.path.exists(XLSX):
        sys.exit("content.xlsx not found next to this repo. Nothing to build.")
    wb = load_workbook(XLSX, data_only=True)
    content = {
        "i18n": build_i18n(wb),
        "services": build_services(wb),
        "locations": build_locations(wb),
        "brands": build_brands(wb),
    }
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(content, f, ensure_ascii=False, indent=2)
    print(f"Wrote {OUT}")
    print(f"  i18n keys : {len(content['i18n'])}")
    print(f"  services  : {len(content['services'])}")
    print(f"  locations : {len(content['locations'])}")
    print(f"  partners  : {len(content['brands'])}")

if __name__ == "__main__":
    main()
